import os
import sqlite3 as sql
import sys
import PyQt5
from math import sqrt
from time import time
import openpyxl
import pandas as pd
from PyQt5 import QtCore
from PyQt5 import uic
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.QtWidgets import QMainWindow, QMenu, QMenuBar, QAction, QFileDialog, QApplication, QTabWidget, QTableView, \
    QShortcut
# ======  Вспомогательные функции ======#
from openpyxl.utils.dataframe import dataframe_to_rows


def M(array):  # рассчет матожидания
    res = sum(array) / len(array)
    return res


def sigma(array):  # расчет дисперсии
    m = M(array)
    return sum([(x - m) ** 2 for x in array])


def isnan(num):
    if num != num:
        return True
    else:
        return False


def remove_nan(arr1, arr2):  # проверка массивов на Nan и их замена на нормальные числа
    if len(arr1) > len(arr2):
        arr1 = arr1[:len(arr2)]
    else:
        arr2 = arr2[:len(arr1)]

    for i, x in enumerate(arr1):
        if isnan(x):
            if isnan(arr2[i]):
                arr1[i] = 0
                arr2[i] = 0
            else:
                arr1[i] = arr2[i]

    for i, x in enumerate(arr2):
        if isnan(x):
            if isnan(arr1[i]):
                arr1[i] = 0
                arr2[i] = 0
            else:
                arr2[i] = arr1[i]

    return arr1, arr2


# ======================================#

class PMOD(QMainWindow):

    def __init__(self):

        super(PMOD, self).__init__()
        self.act2 = QAction('Информация о процессе', self)
        self.refmenu = QMenu("Справка")
        self.act1 = QAction("Сохранть как...", self)
        self.savemenu = QMenu('Сохранить')
        self.menuBar = QMenuBar()
        self.MyCorWind = QMainWindow()
        self.result_table_tv = QTableView()
        self.fname = ''
        self.fname1 = ''

        uic.loadUi("GUI_data_analyzing.ui", self)
        self.setWindowTitle('VKA')
        self.label_db_name.setText('БД не выбранна')
        self.setWindowIcon(QIcon('VKA.jfif'))
        self.label_db_name.setFont(QFont("Arial", 12, QFont.Black))
        self.label_db_name.setAlignment(Qt.AlignCenter)
        self.label_db_name.setStyleSheet("background-color:rgb(150, 150, 200)")
        # self.groupBox.setStyleSheet("background-color:rgb(0,255,0)");
        # self.label_logo.setPixmap(QPixmap("VKA.png"))

        # Подсказки к кнопкам self.btn_get_folder_data.setToolTip('Рекурсивно загружает в БД данныые \nиз всех файлов
        # в указанной директории') self.btn_create_new_db.setToolTip('Создает новую пустую БД')
        self.btn_choose_db.setToolTip('Подключается к уже существущей БД')
        self.comboBox.setToolTip('Все имеющие в БД таблицы \nОтображает выбранную в окне')
        self.spinBox.setToolTip('Количество выводимых строк из таблицы')
        self.spinBox_start_pos.setToolTip('Индекс первой отображаемой строки')
        self.btn_create_corr_table.setToolTip('Создает таблицу корреляции \nВ указанных временных интервалах')
        # self.dspinBox_shift.setToolTip('Временные интервалы обощения параметров')
        self.dspinBox_low_time.setToolTip('Время, начиная с которого будет \nрассчитываться корреляция парметров')
        self.dspinBox_up_time.setToolTip('Время, заканчивая которым будет \nрассчитываться корреляция парметров')
        self.label_info.setToolTip(
            'Здесь отображается информация о \nтекущем состоянии процесса \nсоздания таблицы корреляции')
        # self.radioButton('Режим 1, только на интервале времени , не настраевымый')
        # elf.radioButton_2('Режим 2 , расширенный, настраевымый интервал с зааданным шагом (окном) ')

        # Привязка функций к кнопкам
        # self.btn_get_folder_data.clicked.connect(self.get_folder_data)
        # self.btn_create_new_db.clicked.connect(self.create_new_db)
        self.btn_choose_db.clicked.connect(self.choose_db)
        self.comboBox.currentTextChanged.connect(self.draw_table)
        self.spinBox.valueChanged.connect(self.draw_table)
        self.spinBox_start_pos.valueChanged.connect(self.draw_table)
        self.btn_create_corr_table.clicked.connect(self.create_corr_table)
        self.radioButton_2.clicked.connect(self.mod_2)
        self.radioButton.clicked.connect(self.mod_1)
        self.radioButton.setDisabled(True)
        self.radioButton_2.setDisabled(True)
        self.doubleSpinBox_2.setDisabled(True)
        self.doubleSpinBox_3.setDisabled(True)
        self.doubleSpinBox.setDisabled(True)

        # self.radioButton.clicked.connect(self.choose_db)

        # Включение выключение кнопок
        # self.btn_get_folder_data.setDisabled(True)
        self.comboBox.setDisabled(True)
        self.tableView.setDisabled(True)
        self.spinBox.setDisabled(True)
        self.spinBox_start_pos.setDisabled(True)
        # self.dspinBox_shift.setDisabled(True)
        self.dspinBox_low_time.setDisabled(True)
        self.dspinBox_up_time.setDisabled(True)
        self.progressBar.setDisabled(True)
        self.btn_create_corr_table.setDisabled(True)

        # Объявление необходимых переменных
        self.work_dir = os.getcwd()
        self.tables_list = []
        self.db_name = ''
        self.db_path = ''
        self.conn = None
        self.cursor = None

        self.tabWidget = QTabWidget()
        self.tabWidget.resize(1000, 500)
        self.tabWidget.setWindowTitle('Таблицы корреляции')
        self.tabWidget.setWindowIcon(QIcon('VKA.jfif'))
        self.shortcut = QShortcut(QKeySequence("Ctrl+W"), self.tabWidget, self.close_tab)
        self.tabWidget.tabBar().setMovable(True)
        self.my_massive = []
        self.res = {}

    def choose_db(self):
        print("\n[#] Function: choose_db")

        path, done = QFileDialog.getOpenFileName(self, 'Открытие файла', self.work_dir)
        if done:
            name = path.split('/')[-1]
            print(f"\t {name}")

            try:
                conn = sql.connect(path)
                cursor = conn.cursor()
            except Exception as ex:
                print("\t При открытии возникла ошибка: \n", ex)
                return 0

            self.db_path = path
            self.db_name = path.split('/')[-1]
            self.conn = conn
            self.cursor = cursor

            # Включаем доступность интерфейса
            # self.btn_get_folder_data.setEnabled(True)
            self.comboBox.setEnabled(True)
            self.tableView.setEnabled(True)
            self.spinBox.setEnabled(True)
            self.radioButton.setEnabled(True)
            self.radioButton_2.setEnabled(True)
            self.doubleSpinBox_2.setEnabled(True)
            self.doubleSpinBox_3.setEnabled(True)
            self.doubleSpinBox.setEnabled(True)

            self.label_db_name.setText('Текущая БД:\n {}'.format(name))
            self.label_db_name.setFont(QFont("Arial", 12, QFont.Black))
            self.label_db_name.setAlignment(Qt.AlignCenter)

            # Подгружаем данные в таблицк если они есть
            self.refresh_ui()
        else:
            pass

        return 0

    def refresh_ui(self):
        # заполнение форм
        try:
            cursor = self.cursor
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
            res = [el[0] for el in cursor.fetchall()]

            if len(res) > 1:
                self.tables_list = []
                self.tables_list.extend(res)
                self.tables_list.sort()

                self.comboBox.clear()
                for table_name in res:
                    self.comboBox.addItem(table_name)

                # находим минимальное и максимальное время
                # затем значения помещаем в формы
                first_values = []
                end_values = []

                tables_list = [name for name in self.tables_list if name != 'result_table']
                for table_name in tables_list:
                    cursor.execute("SELECT MIN(time) FROM '{}';".format(table_name))
                    tmp = cursor.fetchone()
                    if tmp[0] is not None:
                        first_values.extend(tmp)

                    cursor.execute("SELECT MAX(time) FROM '{}';".format(table_name))
                    tmp = cursor.fetchone()
                    if tmp[0] is not None:
                        end_values.extend(tmp)

                start_time = min([float(el) for el in first_values])
                end_time = max([float(el) for el in end_values])
                self.doubleSpinBox_2.setRange(start_time, end_time)
                self.doubleSpinBox_3.setRange(start_time, end_time)

                self.dspinBox_low_time.setRange(start_time, end_time)

                self.dspinBox_up_time.setRange(start_time, end_time)

                # self.dspinBox_shift.setEnabled(True)
                self.dspinBox_up_time.setEnabled(True)
                self.dspinBox_low_time.setEnabled(True)
                self.btn_create_corr_table.setEnabled(True)

            else:
                self.comboBox.clear()
                self.comboBox.addItem("В БД нет таблиц")

                # self.dspinBox_shift.setDisabled(True)
                self.dspinBox_low_time.setDisabled(True)
                self.dspinBox_up_time.setDisabled(True)
                self.btn_create_corr_table.setDisabled(True)

        except Exception as ex:
            print('[!] Во вермя заполнения форм возникла ошибка:', ex)

    def draw_table(self):
        print("\n[#] Function: draw_table")

        table_name = self.comboBox.currentText()
        params_number = self.spinBox.value()
        shift = self.spinBox_start_pos.value()

        print('\t Text in comboBox:', [table_name])
        print('\t Number in spinBox:', [params_number])
        print('\t Sift in spinBox_start_pos:', [shift])
        print()
        cursor = self.cursor

        if table_name != 'result_table':
            if table_name in self.tables_list:
                cursor.execute("SELECT * FROM '{}' LIMIT {} OFFSET {};".format(table_name, params_number, shift))
                data = cursor.fetchall()
                print(data)

                # заполняем TableView
                model = QStandardItemModel()

                column_header = ['Время', 'Параметр  "{}"'.format(table_name)]
                row_header = [str(i) for i in range(shift, shift + len(data))]

                for i, line in enumerate(data):
                    time, param_value = line

                    item = QStandardItem(str(time))
                    model.setItem(i, 0, item)

                    item = QStandardItem(str(param_value))
                    model.setItem(i, 1, item)

                model.setHorizontalHeaderLabels(column_header)
                model.setVerticalHeaderLabels(row_header)
                self.tableView.setModel(model)
                self.tableView.horizontalHeader().setStretchLastSection(True)

                self.tableView.setColumnWidth(0, 150)
                self.tableView.setColumnWidth(1, 180)
                self.tableView.setColumnWidth(3, 180)

                self.tableView.setEnabled(True)
                self.spinBox.setEnabled(True)
                self.spinBox_start_pos.setEnabled(True)

            else:
                # чистим поле для таблиц и делаем его недоступным
                self.tableView.setModel(QStandardItemModel())
                self.tableView.setDisabled(True)
                self.spinBox.setDisabled(True)
                self.spinBox_start_pos.setDisabled(True)
        else:
            try:
                # отрисовка для result_table
                tables_list = self.tables_list[:]
                tables_list.sort()
                if 'result_table' in tables_list:
                    tables_list.remove('result_table')

                cursor.execute("SELECT * FROM result_table LIMIT {} OFFSET {};".format(params_number, shift))
                data = cursor.fetchall()

                # заполняем TableView
                model = QStandardItemModel()

                column_header = ['time'] + tables_list
                row_header = [str(i) for i in range(shift, shift + len(data))]

                for i, params_arr in enumerate(data):
                    for j in range(len(params_arr)):
                        item = QStandardItem(str(params_arr[j]))
                        model.setItem(i, j, item)

                model.setHorizontalHeaderLabels(column_header)
                model.setVerticalHeaderLabels(row_header)
                self.result_table_tv.setModel(model)
                self.result_table_tv.horizontalHeader().setStretchLastSection(True)

                self.result_table_tv.setColumnWidth(0, 20)
                for i in range(1, len(column_header)):
                    self.result_table_tv.setColumnWidth(i, 100)

                self.spinBox.setEnabled(True)
                self.spinBox_start_pos.setEnabled(True)

                self.result_table_tv.setWindowTitle('Общая таблица')
                self.result_table_tv.resize(1000, 500)
                self.result_table_tv.show()

            except Exception as ex:
                print('\t\t\t При отображении result_table возника ошибка:\n\t\t ', ex)
                return 0

        return 0

    def create_corr_table(self):
        if self.radioButton.isChecked():
            try:
                t_start = time()

                cursor = self.cursor
                # step =  round(self.dspinBox_shift.value(), 3)
                start_time = round(self.dspinBox_low_time.value(), 3)
                end_time = round(self.dspinBox_up_time.value(), 3)
                tables_list = [name for name in self.tables_list if name != 'result_table']

                if end_time < start_time:
                    end_time, start_time = start_time, end_time

                print('[#] Function: create_corr_table')
                # print('\t Step: ', step)
                print('\t Start_time: ', start_time)
                print('\t End_time: ', end_time)

                # вытаскиваем данные из БД
                print('\n\t [+] Colecting data from DB:')
                self.label_info.setText('Загрузка данных из БД')

                # получаем значения в предлах указанных границ
                data = []
                percent = len(tables_list) * 0.01
                params_count = len(tables_list)
                for i, table_name in enumerate(tables_list):
                    if round(i % percent, 0) == 0:
                        print(f'\t\t => {i // percent}% \tcurrent tale {i} from {params_count}')
                        self.progressBar.setValue(i / percent)

                    cursor.execute(
                        "SELECT param FROM '{}' WHERE time >= {} AND time <= {};".format(table_name, start_time,
                                                                                         end_time))
                    tmp = [el[0] for el in cursor.fetchall()]
                    print(table_name)
                    print(1)
                    print(data)
                    print(tmp)

                    if len(tmp) == 0:
                        data.append([float('NaN')])
                    else:
                        data.append([float(x) for x in tmp])
                    print(data)

                # считаем корреляцию
                print('\n\t [+] Computing correlations:')
                self.label_info.setText('Рассчет корреляции')

                params_count = len(data)
                percent = params_count * 0.01
                corr_table = [[0 for _ in range(params_count)] for _ in range(params_count)]

                for j in range(params_count):
                    if round(j % percent, 0) == 0:
                        print(f'\t\t => {j // percent}% \tcurrent line {j} from {params_count}')
                        self.progressBar.setValue(j / percent)

                    for i in range(params_count):
                        if i != j:
                            # проверка массивов на наличеие NaN
                            arr1, arr2 = remove_nan(data[i][:], data[j][:])
                            M_p1, M_p2 = M(arr1), M(arr2)

                            up = sum((x - M_p1) * (y - M_p2) for x, y in zip(arr1, arr2))
                            down = sqrt(sigma(arr1) * sigma(arr2))
                            if down == 0:
                                r = 0
                            else:
                                r = up / down
                            corr_table[i][j] = round(r, 3)
                            self.my_massive = corr_table
                print(self.my_massive)

                print('\n\t [+] Drawing table:')
                self.label_info.setText('Создание таблицы')

                # заполняем TableView
                model = QStandardItemModel()

                # column_header = [str(round(start_time + i*step, 2)) for i in range(int(abs(end_time -
                # start_time)/step) + 1)] row_header = column_header[:]

                column_header = [name for name in self.tables_list if name != 'result_table']
                row_header = column_header[:]

                for j in range(params_count):
                    if round(j % percent, 0) == 0:
                        print(f'\t\t => {j // percent}% \tcurrent line {j} from {params_count}')

                        self.progressBar.setValue(j / percent)

                    for i in range(params_count):
                        if i != j:
                            value = corr_table[i][j]
                            item = QStandardItem(str(value))

                            # задаем цвет
                            if value > 0:
                                r = int(255 * abs(1 - value))
                                g = 255
                                b = int(255 * abs(1 - value))
                            else:
                                r = 255
                                g = int(255 * abs(1 + value))
                                b = int(255 * abs(1 + value))

                            model.setItem(i, j, item)
                            model.setData(model.index(i, j), QBrush(QColor(r, g, b)), QtCore.Qt.BackgroundRole)
                        else:
                            item = QStandardItem('-')
                            model.setItem(i, j, item)

                model.setHorizontalHeaderLabels(column_header)
                model.setVerticalHeaderLabels(row_header)

                corr_tableView = QTableView()
                corr_tableView.setModel(model)
                # corr_tableView.setWindowTitle('Общая таблица')
                # corr_tableView.setWindowIcon(QIcon('VKA.jfif'))

                for i in range(params_count):
                    corr_tableView.setColumnWidth(i, 10)

                # self.corr_tableView.show()
                st, et = round(start_time, 2), round(end_time, 2)

                self.tabWidget.addTab(corr_tableView, '{} to {}'.format(st, et))
                # self.tabWidget.show()
                ###############################################
                # Создаю экземпляр класса#######
                self.MyCorWind.setWindowTitle('Таблица корреляции')

                # Размещаю в центральном виджете виджет таблицы
                self.MyCorWind.setCentralWidget(self.tabWidget)
                # Создаю обьект менюбар
                # До бавляю меню бар в мое главное окно и устанавливаю пункты меню и указываю подразделы
                self.MyCorWind.setMenuBar(self.menuBar)

                self.menuBar.addMenu(self.savemenu)
                self.savemenu.addAction(self.act1)

                self.refmenu.addAction(self.act2)
                self.menuBar.addMenu(self.refmenu)
                self.act1.triggered.connect(self.save_on)
                self.act2.triggered.connect(self.save_on)

                self.MyCorWind.show()

                t_end = time()

                print('Spend time: {} sec \n'.format(t_end - t_start))
                self.progressBar.setValue(0)
                self.label_info.setText('Выполненно')
            except Exception as ex:
                print('\t [!] Во время создания таблицы корреляции возникла ошибка:', ex)
        if self.radioButton_2.isChecked():
            wind_step = round(self.doubleSpinBox_3.value(), 3)
            start_time = round(self.doubleSpinBox.value(), 3)
            end_time = round(self.doubleSpinBox_2.value(), 3)
            print(start_time)
            print(end_time)
            print(wind_step)
            start_time_1 = start_time
            end_time_1 = start_time_1 + wind_step
            self.tables_list.remove('result_table')
            fname, _ = QFileDialog.getSaveFileName(self, 'Save', self.fname, "*.xlsx")
            if not fname: return
            self.fname = fname

            while end_time_1 <= end_time:

                print(404)
                try:
                    t_start = time()

                    cursor = self.cursor
                    # step =  round(self.dspinBox_shift.value(), 3)
                    start_time = round(self.doubleSpinBox.value(), 3)
                    tables_list = [name for name in self.tables_list if name != 'result_table']

                    if end_time < start_time:
                        end_time, start_time = start_time, end_time

                    print('[#] Function: create_corr_table')
                    # print('\t Step: ', step)
                    print('\t Start_time: ', start_time_1)
                    print('\t End_time_1: ', end_time_1)
                    print('\t Wind_step: ', wind_step)

                    # вытаскиваем данные из БД
                    print('\n\t [+] Colecting data from DB:')
                    self.label_info.setText('Загрузка данных из БД')

                    # получаем значения в предлах указанных границ
                    data = []
                    percent = len(tables_list) * 0.01
                    params_count = len(tables_list)
                    for i, table_name in enumerate(tables_list):
                        if round(i % percent, 0) == 0:
                            print(f'\t\t => {i // percent}% \tcurrent tale {i} from {params_count}')
                            self.progressBar.setValue(i / percent)

                        cursor.execute(
                            "SELECT param FROM '{}' WHERE time >= {} AND time <= {};".format(table_name, start_time_1,
                                                                                             end_time_1))
                        tmp = [el[0] for el in cursor.fetchall()]

                        print(table_name)
                        print(1)
                        print(data)
                        print(tmp)

                        if len(tmp) == 0:
                            data.append([float('NaN')])
                        else:
                            data.append([float(x) for x in tmp])
                        print(data)

                    # считаем корреляцию
                    print('\n\t [+] Computing correlations:')
                    self.label_info.setText('Рассчет корреляции')

                    params_count = len(data)
                    percent = params_count * 0.01
                    corr_table = [[0 for _ in range(params_count)] for _ in range(params_count)]

                    for j in range(params_count):
                        if round(j % percent, 0) == 0:
                            print(f'\t\t => {j // percent}% \tcurrent line {j} from {params_count}')
                            self.progressBar.setValue(j / percent)

                        for i in range(params_count):
                            if i != j:
                                # проверка массивов на наличеие NaN
                                arr1, arr2 = remove_nan(data[i][:], data[j][:])
                                M_p1, M_p2 = M(arr1), M(arr2)

                                up = sum((x - M_p1) * (y - M_p2) for x, y in zip(arr1, arr2))
                                down = sqrt(sigma(arr1) * sigma(arr2))
                                if down == 0:
                                    r = 0
                                else:
                                    r = up / down
                                corr_table[i][j] = round(r, 3)
                                self.my_massive = corr_table
                    print(self.my_massive)

                    print('\n\t [+] Drawing table:')
                    self.label_info.setText('Создание таблицы')

                    # заполняем TableView
                    model = QStandardItemModel()

                    column_header = [name for name in self.tables_list if name != 'result_table']
                    row_header = column_header[:]

                    for j in range(params_count):
                        if round(j % percent, 0) == 0:
                            print(f'\t\t => {j // percent}% \tcurrent line {j} from {params_count}')

                            self.progressBar.setValue(j / percent)

                        for i in range(params_count):
                            if i != j:
                                value = corr_table[i][j]
                                item = QStandardItem(str(value))

                                # задаем цвет
                                if value > 0:
                                    r = int(255 * abs(1 - value))
                                    g = 255
                                    b = int(255 * abs(1 - value))
                                else:
                                    r = 255
                                    g = int(255 * abs(1 + value))
                                    b = int(255 * abs(1 + value))

                                model.setItem(i, j, item)
                                model.setData(model.index(i, j), QBrush(QColor(r, g, b)), QtCore.Qt.BackgroundRole)
                            else:
                                item = QStandardItem('-')
                                model.setItem(i, j, item)

                    model.setHorizontalHeaderLabels(column_header)
                    model.setVerticalHeaderLabels(row_header)

                    corr_tableView = QTableView()
                    corr_tableView.setModel(model)
                    # corr_tableView.setWindowTitle('Общая таблица')
                    # corr_tableView.setWindowIcon(QIcon('VKA.jfif'))

                    for i in range(params_count):
                        corr_tableView.setColumnWidth(i, 10)

                    # self.corr_tableView.show()
                    st, et = round(start_time_1, 2), round(end_time_1, 2)

                    self.tabWidget.addTab(corr_tableView, '{} to {}'.format(st, et))

                    # self.tabWidget.show()
                    ##############################################################
                    # Создаю экземпляр класса#######
                    self.MyCorWind.setWindowTitle('Таблица корреляции')

                    # Размещаю в центральном виджете виджет таблицы
                    self.MyCorWind.setCentralWidget(self.tabWidget)
                    # Создаю обьект менюбар
                    # Добавляю меню бар в мое главное окно и устанавливаю пункты меню и указываю подразделы
                    self.MyCorWind.setMenuBar(self.menuBar)

                    self.menuBar.addMenu(self.savemenu)
                    self.savemenu.addAction(self.act1)

                    self.refmenu.addAction(self.act2)
                    self.menuBar.addMenu(self.refmenu)
                    self.act1.triggered.connect(self.save_on)
                    self.act2.triggered.connect(self.save_on)

                    self.MyCorWind.show()

                    t_end = time()

                    print('Spend time: {} sec \n'.format(t_end - t_start))
                    self.progressBar.setValue(0)
                    self.label_info.setText('Выполненно')

                    start_time_1 = wind_step + start_time_1
                    end_time_1 = wind_step + end_time_1
                    print(111111)

                    self.res = {self.tables_list[i]: self.my_massive[i] for i in range(len(self.tables_list))}
                    # Создаем фрейм - таблицу  из полученных словарей.
                    data_frame = pd.DataFrame(self.res)
                    # Добавляем колонку None для удобного восприятия результатов.
                    data_frame[None] = self.tables_list
                    # Получаем доступ к колонкам фрейма для того , чтобы перенести новый столбец в начало
                    cols = data_frame.columns.tolist()
                    # Перемещаем стобец None из конца в начало.
                    cols = cols[-1:] + cols[:-1]
                    data_frame = data_frame[cols]
                    print(data_frame)

                    try:
                        wb = openpyxl.load_workbook(self.fname)
                    except:
                        wb = openpyxl.Workbook()
                        # Удаление листа, создаваемого по умолчанию, при создании документа
                        for sheet_name in wb.sheetnames:
                            sheet = wb.get_sheet_by_name(sheet_name)
                            wb.remove_sheet(sheet)
                    # Создаю новую страницу
                    ws = wb.create_sheet('Таблица')
                    # Записываю фрейм
                    for r in dataframe_to_rows(data_frame, index=False, header=True):
                        ws.append(r)
                    # Сохраняю
                    from openpyxl.writer.excel import save_workbook
                    save_workbook(wb, self.fname)

                except Exception as ex:
                    print('\t [!] Во время создания таблицы корреляции возникла ошибка:', ex)

    def save_on(self):
        self.tables_list.remove('result_table')
        fname, _ = QFileDialog.getSaveFileName(self, 'Save', self.fname, "*.xls")
        if not fname: return
        self.fname = fname
        # Создаем словари из двух списков , для создания таблицы результатов.
        self.res = {self.tables_list[i]: self.my_massive[i] for i in range(len(self.tables_list))}
        # Создаем фрейм - таблицу  из полученных словарей.
        data_frame = pd.DataFrame(self.res)
        # Добавляем колонку None для удобного восприятия результатов.
        data_frame[None] = self.tables_list
        # Получаем доступ к колонкам фрейма для того , чтобы перенести новый столбец в начало
        cols = data_frame.columns.tolist()
        # Перемещаем стобец None из конца в начало.
        cols = cols[-1:] + cols[:-1]
        data_frame = data_frame[cols]
        print(data_frame)
        data_frame.to_excel(self.fname, index=False)

    def close_tab(self):
        try:
            if self.table.count():
                self.table.removeTab(self.table.currentIndex())
                print(f"[!] Tab {self.table.currentIndex()} closed")

                if self.table.count() == 0:
                    self.table.close()

        except Exception as ex:
            print('\t [!] Во время удаления таблицы возникла ошибка:', ex)

    def mod_1(self):
        print(57)
        self.doubleSpinBox_2.setDisabled(True)
        self.doubleSpinBox_3.setDisabled(True)
        self.doubleSpinBox.setDisabled(True)
        self.dspinBox_up_time.setEnabled(True)
        self.dspinBox_low_time.setEnabled(True)

    def mod_2(self):
        print(77)
        self.doubleSpinBox_2.setEnabled(True)
        self.doubleSpinBox_3.setEnabled(True)
        self.doubleSpinBox.setEnabled(True)
        self.dspinBox_up_time.setDisabled(True)
        self.dspinBox_low_time.setDisabled(True)


app = QApplication(sys.argv)
ex = PMOD()
ex.show()
sys.exit(app.exec_())
